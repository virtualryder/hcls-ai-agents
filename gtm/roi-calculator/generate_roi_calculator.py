#!/usr/bin/env python3
"""
HCLS AI Agent Suite — ROI Calculator generator.

Produces HCLS-AI-Suite-ROI-Calculator.xlsx: an SA-fillable workbook that turns the
cited cost-of-doing-nothing anchors in ../HCLS-DECK-SOURCES.md into a per-agent,
customer-specific ROI model. Blue cells are inputs the SA edits; everything else is
a formula. Figures are modeled at the customer's own baseline and are not guaranteed.

Run:  python3 gtm/roi-calculator/generate_roi_calculator.py
Deps: pip install openpyxl --break-system-packages
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "232F3E"; ORANGE = "FF9900"; TEAL = "16A085"; GRAY = "F2F3F4"
INPUT = "DCE9F7"; WHITE = "FFFFFF"
thin = Side(style="thin", color="D5DBE1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, text, row=1, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 30

def subhdr(ws, text, row, span=8, color=ORANGE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=11, color=NAVY)
    c.fill = PatternFill("solid", fgColor=GRAY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def labelcell(ws, r, c, v, bold=False, italic=False, wrap=False, size=10):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, italic=italic, size=size, color=NAVY)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    return cell

def inputcell(ws, r, c, v, fmt="#,##0"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = PatternFill("solid", fgColor=INPUT)
    cell.font = Font(bold=True, color="1F4E79")
    cell.number_format = fmt
    cell.border = border
    cell.alignment = Alignment(horizontal="right")
    return cell

def formulacell(ws, r, c, f, fmt="$#,##0", bold=False):
    cell = ws.cell(row=r, column=c, value=f)
    cell.number_format = fmt
    cell.font = Font(bold=bold, color=TEAL if bold else NAVY)
    cell.border = border
    cell.alignment = Alignment(horizontal="right")
    return cell

wb = Workbook()

# ---------------- README / how-to sheet ----------------
ws = wb.active; ws.title = "Read me"
ws.column_dimensions["A"].width = 110
hdr(ws, "HCLS AI Agent Suite — ROI Calculator", span=1)
notes = [
    "",
    "HOW TO USE",
    "1. Open the 'Assumptions' tab and set the blue input cells to your organization's baseline.",
    "2. Open each agent tab; set the blue per-agent input cells (volumes, rates, expected reduction).",
    "3. Read the green 'Modeled annual value' and the 'Summary' tab. Edit only blue cells.",
    "",
    "DISCIPLINE (so the model stays board-defensible)",
    "• Every anchor figure traces to ../HCLS-DECK-SOURCES.md with a source-class tag.",
    "• Outcomes are modeled at YOUR baseline — they are not guaranteed and depend on data quality,",
    "  process design, and adoption.",
    "• 'Expected reduction %' is a planning assumption the SA sets WITH the customer, not a vendor promise.",
    "• Bedrock inference + infrastructure + SI delivery cost must be modeled separately (see Assumptions).",
    "",
    "SOURCE ANCHORS USED",
    "• Clinical-trial delay: ~$800K/day lost sales + Phase III $55,716/day direct — Tufts CSDD 2024.",
    "• Drug development cost: $2.6B per approved drug — DiMasi/Tufts 2016.",
    "• Submission delay: ~$60M NPV/month for a $1B-peak asset — McKinsey 2025.",
    "• Protocol amendment: ~$535K direct (Phase III), ~45% avoidable — Tufts CSDD 2016.",
    "• PV intake: 40–85% of PV budget; ICSR entry time cut 40–70% — Schmider 2019.",
    "• RWE: ~45% of analyst time on data prep — Anaconda.",
    "• MLR: 50–70% time-to-deliver reduction potential — McKinsey.",
]
for i, t in enumerate(notes, start=2):
    c = ws.cell(row=i, column=1, value=t)
    if t in ("HOW TO USE", "DISCIPLINE (so the model stays board-defensible)", "SOURCE ANCHORS USED"):
        c.font = Font(bold=True, size=12, color=NAVY)
    else:
        c.font = Font(size=10, color=NAVY)

# ---------------- Assumptions ----------------
wa = wb.create_sheet("Assumptions")
for col, w in [("A", 48), ("B", 16), ("C", 60)]:
    wa.column_dimensions[col].width = w
hdr(wa, "Global Assumptions  (edit blue cells)", span=3)
labelcell(wa, 2, 1, "Parameter", bold=True); labelcell(wa, 2, 2, "Value", bold=True); labelcell(wa, 2, 3, "Note", bold=True)
ASSUMPTIONS = [
    ("loaded_fte", "Fully-loaded labor cost / FTE / yr", 175000, "$#,##0", "Blended loaded rate; adjust per geography/role."),
    ("loaded_hr", "Fully-loaded labor cost / hour", 95, "$#,##0", "Used for hours-saved conversions."),
    ("delay_day_sales", "Lost/delayed peak sales per day of delay", 800000, "$#,##0", "Tufts CSDD 2024 (asset-dependent)."),
    ("ph3_direct_day", "Phase III direct trial cost per day", 55716, "$#,##0", "Tufts CSDD 2024."),
    ("amend_ph3", "Direct cost of a substantial Phase III amendment", 535000, "$#,##0", "Tufts CSDD 2016."),
    ("bedrock_year", "Estimated Bedrock inference + infra / yr (per agent)", 120000, "$#,##0", "Replace with your TCO-model figure."),
]
r = 3
keys = {}
for key, label, val, fmt, note in ASSUMPTIONS:
    labelcell(wa, r, 1, label)
    inputcell(wa, r, 2, val, fmt)
    labelcell(wa, r, 3, note, italic=True, size=9, wrap=True)
    keys[key] = f"Assumptions!$B${r}"
    r += 1

# ---------------- Per-agent sheets ----------------
# Each agent: a small driver model. (input rows) -> cost of doing nothing + modeled value.
AGENTS = [
    {
        "id": "01", "name": "Regulatory Writing",
        "drivers": [
            ("Submissions / major filings per year", 6, "#,##0"),
            ("Avg. writer hours per submission (drafting+assembly)", 1200, "#,##0"),
            ("Expected hours reduction % (Merck-class pilots ~40-55%)", 0.40, "0%"),
            ("Avg. submission delay avoided (days/filing)", 10, "#,##0"),
        ],
        # value = labor saved + delay-days avoided * sales/day
        "value": lambda a, k: f"=(B3*B4*B5*{k['loaded_hr']}) + (B3*B6*{k['delay_day_sales']})",
        "cdn": lambda a, k: f"=B3*B6*{k['delay_day_sales']}",
        "cdn_label": "Cost of doing nothing (delay NPV exposure, modeled)",
    },
    {
        "id": "02", "name": "Pharmacovigilance",
        "drivers": [
            ("ICSR cases processed per year", 100000, "#,##0"),
            ("Avg. processing minutes per case today", 69, "#,##0"),
            ("Expected processing-time reduction % (Schmider 40-70%)", 0.40, "0%"),
            ("Blended cost per processing hour", 95, "$#,##0"),
        ],
        "value": lambda a, k: "=(B3*B4/60)*B5*B6",
        "cdn": lambda a, k: "=(B3*B4/60)*B6",
        "cdn_label": "Annual case-processing labor today (modeled)",
    },
    {
        "id": "03", "name": "Clinical Trial Ops & TMF",
        "drivers": [
            ("Active pivotal (Phase III) trials", 3, "#,##0"),
            ("Avg. database-lock / TMF slip avoided (days/trial)", 20, "#,##0"),
            ("Lost sales per day of delay", 800000, "$#,##0"),
            ("Phase III direct cost per day", 55716, "$#,##0"),
        ],
        "value": lambda a, k: "=B3*B4*(B5+B6)",
        "cdn": lambda a, k: "=B3*B4*(B5+B6)",
        "cdn_label": "Exposure per avoided slip across active trials (modeled)",
    },
    {
        "id": "04", "name": "Site & Patient Matching",
        "drivers": [
            ("New trial launches per year", 4, "#,##0"),
            ("Enrollment days pulled in per launch", 30, "#,##0"),
            ("Lost sales per day of delay", 800000, "$#,##0"),
            ("Non-enrolling site activations avoided / yr", 8, "#,##0"),
            ("Cost per wasted site activation", 40000, "$#,##0"),
        ],
        "value": lambda a, k: "=(B3*B4*B5)+(B6*B7)",
        "cdn": lambda a, k: "=B3*B4*B5",
        "cdn_label": "Accelerated revenue from faster enrollment (modeled)",
    },
    {
        "id": "05", "name": "Quality / CAPA",
        "drivers": [
            ("Complaints / quality events per year", 5000, "#,##0"),
            ("Avg. analyst hours per CAPA / investigation", 12, "#,##0"),
            ("Expected hours reduction % (triage+RCA draft)", 0.30, "0%"),
            ("Blended cost per hour", 95, "$#,##0"),
            ("Recall-exposure value at risk reduced / yr", 0, "$#,##0"),
        ],
        "value": lambda a, k: "=(B3*B4*B5*B6)+B7",
        "cdn": lambda a, k: "=B3*B4*B6",
        "cdn_label": "Annual CAPA / investigation labor today (modeled)",
    },
    {
        "id": "06", "name": "Protocol Design",
        "drivers": [
            ("Phase III protocols authored per year", 4, "#,##0"),
            ("Avoidable substantial amendments prevented / protocol", 0.5, "0.0"),
            ("Direct cost per Phase III amendment", 535000, "$#,##0"),
            ("Cycle delay avoided per prevented amendment (days)", 70, "#,##0"),
            ("Lost sales per day of delay", 800000, "$#,##0"),
        ],
        "value": lambda a, k: "=B3*B4*(B5+(B6*B7))",
        "cdn": lambda a, k: "=B3*B4*B5",
        "cdn_label": "Direct amendment cost avoided (modeled)",
    },
    {
        "id": "07", "name": "RWE / HEOR",
        "drivers": [
            ("RWE / HEOR analysts (FTE)", 20, "#,##0"),
            ("Fully-loaded cost per analyst / yr", 150000, "$#,##0"),
            ("Share of time on data prep today (Anaconda ~45%)", 0.45, "0%"),
            ("Share of prep time reclaimed", 0.50, "0%"),
        ],
        "value": lambda a, k: "=B3*B4*B5*B6",
        "cdn": lambda a, k: "=B3*B4*B5",
        "cdn_label": "Non-analytic labor cost today (modeled)",
    },
    {
        "id": "08", "name": "Medical Affairs / MSL",
        "drivers": [
            ("MLR-reviewed content items per year", 4000, "#,##0"),
            ("Avg. reviewer hours per item", 3, "#,##0"),
            ("Expected time-to-deliver reduction % (McKinsey 50-70%)", 0.50, "0%"),
            ("Blended cost per reviewer hour", 110, "$#,##0"),
        ],
        "value": lambda a, k: "=B3*B4*B5*B6",
        "cdn": lambda a, k: "=B3*B4*B6",
        "cdn_label": "Annual MLR review labor today (modeled)",
    },
]

summary_rows = []
for a in AGENTS:
    tab_title = f"{a['id']} {a['name']}".replace("/", "-")[:31]
    ws = wb.create_sheet(tab_title)
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 18
    hdr(ws, f"Agent {a['id']} — {a['name']}  (edit blue cells)", span=2)
    subhdr(ws, "Inputs (your baseline)", 2, span=2)
    for i, (label, val, fmt) in enumerate(a["drivers"]):
        rr = 3 + i
        labelcell(ws, rr, 1, label, wrap=True)
        inputcell(ws, rr, 2, val, fmt)
    last = 2 + len(a["drivers"])
    # cost of doing nothing
    cdn_row = last + 2
    subhdr(ws, "Results (modeled)", cdn_row - 1, span=2)
    labelcell(ws, cdn_row, 1, a["cdn_label"], bold=True)
    formulacell(ws, cdn_row, 2, a["cdn"](a, keys), bold=False)
    # modeled annual value
    val_row = cdn_row + 1
    labelcell(ws, val_row, 1, "Modeled annual value of the agent", bold=True)
    formulacell(ws, val_row, 2, a["value"](a, keys), bold=True)
    # less platform cost
    cost_row = val_row + 1
    labelcell(ws, cost_row, 1, "Less: est. Bedrock inference + infra / yr", bold=False)
    formulacell(ws, cost_row, 2, f"=-{keys['bedrock_year']}")
    net_row = cost_row + 1
    labelcell(ws, net_row, 1, "Net modeled annual value", bold=True)
    nc = formulacell(ws, net_row, 2, f"=B{val_row}+B{cost_row}", bold=True)
    nc.fill = PatternFill("solid", fgColor="E8F6F1")
    # caveat
    cav = ws.cell(row=net_row + 2, column=1,
                  value="Modeled at your baseline; not guaranteed. SI delivery cost is separate (see TCO model).")
    cav.font = Font(italic=True, size=9, color="6B7785")
    summary_rows.append((a["id"], a["name"], ws.title, val_row, net_row))

# ---------------- Summary ----------------
wsum = wb.create_sheet("Summary", 2)
for col, w in [("A", 8), ("B", 34), ("C", 24), ("D", 24)]:
    wsum.column_dimensions[col].width = w
hdr(wsum, "Suite ROI Summary  (modeled at your baseline)", span=4)
labelcell(wsum, 2, 1, "#", bold=True); labelcell(wsum, 2, 2, "Agent", bold=True)
labelcell(wsum, 2, 3, "Modeled annual value", bold=True); labelcell(wsum, 2, 4, "Net of platform cost", bold=True)
for i, (aid, name, sheet, vrow, nrow) in enumerate(summary_rows):
    rr = 3 + i
    labelcell(wsum, rr, 1, aid)
    labelcell(wsum, rr, 2, name)
    formulacell(wsum, rr, 3, f"='{sheet}'!B{vrow}")
    formulacell(wsum, rr, 4, f"='{sheet}'!B{nrow}")
tot = 3 + len(summary_rows)
labelcell(wsum, tot, 2, "SUITE TOTAL", bold=True)
formulacell(wsum, tot, 3, f"=SUM(C3:C{tot-1})", bold=True)
formulacell(wsum, tot, 4, f"=SUM(D3:D{tot-1})", bold=True)
note = wsum.cell(row=tot + 2, column=1,
                 value="All figures modeled from the cited anchors in HCLS-DECK-SOURCES.md, applied to your inputs. "
                       "Not a guarantee. Land-first agents 01/02/03/08 typically anchor the Phase-1 business case.")
note.font = Font(italic=True, size=9, color="6B7785")
wsum.merge_cells(start_row=tot + 2, start_column=1, end_row=tot + 2, end_column=4)
OUT = os.path.join(os.path.dirname(__file__), "HCLS-AI-Suite-ROI-Calculator.xlsx")
wb.save(OUT)
print("wrote", OUT)
